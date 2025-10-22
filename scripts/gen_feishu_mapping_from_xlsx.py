#!/usr/bin/env python3
import argparse, os, time, sys
import openpyxl, yaml

def detect_columns(sheet):
    header_row = 2  # Feishu export template header row
    headers = {}
    for col in range(1, sheet.max_column + 1):
        v = sheet.cell(header_row, col).value
        if isinstance(v, str):
            headers[v.strip()] = col
    user_col = headers.get('用户 ID') or headers.get('用户ID') or headers.get('employee_id')
    email_col = headers.get('企业邮箱') or headers.get('工作邮箱') or headers.get('email')
    if not user_col or not email_col:
        raise RuntimeError('无法定位列: 需要“用户 ID”和“企业邮箱/工作邮箱”')
    if '企业邮箱' in headers:
        email_col = headers['企业邮箱']
    return header_row, user_col, email_col

def build_mapping(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = wb.active
    header_row, user_col, email_col = detect_columns(sheet)
    mapping = {}
    for row in range(header_row + 1, sheet.max_row + 1):
        uid = sheet.cell(row, user_col).value
        eml = sheet.cell(row, email_col).value
        if uid is None or eml is None:
            continue
        uid = str(uid).strip(); eml = str(eml).strip()
        if uid and eml:
            mapping[uid] = eml
    return mapping

def main():
    ap = argparse.ArgumentParser(description='Generate Feishu user_id → enterprise_email mapping from XLSX export')
    ap.add_argument('--input', required=True, help='Path to Feishu directory export .xlsx')
    ap.add_argument('--output', default='mappings/feishu_userid_to_enterprise_email.yaml', help='Output YAML path')
    args = ap.parse_args()
    mapping = build_mapping(args.input)
    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    out = {
        'generated_at': int(time.time()),
        'source': 'xlsx import',
        'count': len(mapping),
        'mapping': mapping,
    }
    with open(args.output, 'w', encoding='utf-8') as f:
        yaml.safe_dump(out, f, allow_unicode=True, sort_keys=True)
    print(f'Wrote {len(mapping)} entries to {args.output}')

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        print(f'ERROR: {e}', file=sys.stderr)
        sys.exit(1)
